# -*- coding: utf-8 -*-
import os, re, math, unicodedata, datetime as dt, zipfile
from io import BytesIO
from dataclasses import dataclass
from typing import Optional, Tuple, List, Dict, Set
from functools import lru_cache

import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont, ImageFilter

# =========================
# UI
# =========================
st.set_page_config(page_title="Tách bảng lương -> Ảnh (5 lao động/ảnh)", layout="wide")
st.title("Tách bảng lương Excel → Ảnh PNG (5 lao động/ảnh) – Xuất ZIP")
st.caption("Không cắt chữ (không …). Tự wrap + tự tăng chiều cao hàng để hiển thị đủ nội dung trong ô.")

# =========================
# CONFIG
# =========================
@dataclass
class RenderConfig:
    rows_per_image: int = 5
    final_scale: int = 3
    super_sample: int = 2

    zebra_enabled: bool = True
    zebra_a: str = "#ffffff"
    zebra_b: str = "#f3f8ff"

    hide_empty_columns: bool = True
    treat_zero_as_empty: bool = True

    title_text: str = "HỢP CHÍ - Nguyễn Huệ HR"
    footer_text: str = "Liên hệ Nguyễn Huệ HR - 0356 227 868 | timvieclam.9phut.com"

    text_color: str = "#0b1220"
    border_color: str = "#cbd5e1"
    header_bg: str = "#eaf2ff"
    header2_bg: str = "#f3f8ff"

    data_use_bold: bool = True
    data_stroke_width: int = 2
    header_stroke_width: int = 1

    draw_header_divider: bool = True
    header_divider_color: str = "#bbd3ff"

    # NEW: auto-fit text
    enable_wrap: bool = True
    line_gap_px: int = 2         # sẽ nhân theo render_scale
    cell_pad_x_px: int = 6       # sẽ nhân theo render_scale
    cell_pad_y_px: int = 4       # sẽ nhân theo render_scale

    # NEW: auto widen columns (giúp header không bị cụt)
    auto_widen_columns: bool = True
    col_widen_cap_px: int = 520  # cap (px ở final, sẽ nhân theo render_scale)


# =========================
# HEADER DETECT
# =========================
HEADER_SYNONYMS = {
    "stt": {"stt", "s tt", "s tt.", "số tt", "so tt", "no", "no.", "tt"},
    "ma_nv": {"mã nv", "ma nv", "mã nhân viên", "ma nhan vien", "msnv", "employee id", "emp id", "id"},
    "tong_luong": {"tổng lương", "tong luong", "total salary", "tổng tiền lương", "tong tien luong"},
}

def _norm_text(x) -> str:
    if x is None:
        return ""
    s = unicodedata.normalize("NFKC", str(x)).strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s

def _matches_any(token: str, candidates: Set[str]) -> bool:
    if not token:
        return False
    return any(token == k or k in token for k in candidates)

def find_header(ws, scan_rows: int = 120) -> Tuple[int, int, int, int, int]:
    max_col, max_row = ws.max_column, ws.max_row
    for r in range(1, min(scan_rows, max_row) + 1):
        row_map = {
            c: _norm_text(ws.cell(r, c).value)
            for c in range(1, max_col + 1)
            if ws.cell(r, c).value is not None
        }
        if not row_map:
            continue

        col_stt = next((c for c, t in row_map.items() if _matches_any(t, HEADER_SYNONYMS["stt"])), None)
        col_ma  = next((c for c, t in row_map.items() if _matches_any(t, HEADER_SYNONYMS["ma_nv"])), None)
        col_tl  = next((c for c, t in row_map.items() if _matches_any(t, HEADER_SYNONYMS["tong_luong"])), None)

        if col_stt and col_ma and col_tl and col_stt < col_tl:
            header_rows_count = 1
            if r + 1 <= max_row:
                nxt = [_norm_text(ws.cell(r + 1, c).value) for c in range(col_stt, col_tl + 1)]
                if any(("thành tiền" in t or "thanh tien" in t or "số giờ" in t or "so gio" in t) for t in nxt):
                    header_rows_count = 2
            return r, header_rows_count, col_stt, col_ma, col_tl

    raise ValueError("Không tìm thấy hàng tiêu đề (STT / Mã NV / Tổng lương).")

def find_total_row(ws, data_start_row: int, col_stt: int, c_left_scan: int, c_right_scan: int) -> int:
    max_row = ws.max_row
    for r in range(data_start_row, max_row + 1):
        t = _norm_text(ws.cell(r, col_stt).value)
        if t in {"tổng", "tong", "tổng cộng", "tong cong"} or t.startswith("tổng "):
            return r
    for r in range(data_start_row, max_row + 1):
        if any(_norm_text(ws.cell(r, c).value) in {"tổng", "tong", "tổng cộng", "tong cong"} for c in range(c_left_scan, c_right_scan + 1)):
            return r
    raise ValueError("Không tìm thấy dòng 'TỔNG' để dừng tách.")


# =========================
# FORMAT / EMPTY
# =========================
def format_cell_value(cell, thousand_sep: str=".", decimal_sep: str=",") -> str:
    v = cell.value
    if v is None:
        return ""
    if isinstance(v, str):
        return v.replace("\r", " ").replace("\n", " ").strip()
    if isinstance(v, (dt.datetime, dt.date)):
        if isinstance(v, dt.datetime) and v.time() != dt.time(0, 0, 0):
            return v.strftime("%d/%m/%Y %H:%M")
        return v.strftime("%d/%m/%Y")
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, (int, float)):
        fmt = (cell.number_format or "").lower()
        if "%" in fmt:
            s = f"{float(v) * 100.0:.2f}%"
            return s.replace(".", decimal_sep)
        if isinstance(v, float) and abs(v - round(v)) < 1e-9:
            v = int(round(v))
        if isinstance(v, int):
            return f"{v:,}".replace(",", thousand_sep)
        decimals = 2
        m = re.search(r"\.(0+)", fmt)
        if m:
            decimals = len(m.group(1))
        s = f"{float(v):,.{decimals}f}"
        s = s.replace(",", "X").replace(".", decimal_sep).replace("X", thousand_sep)
        return s
    return str(v)

def is_effectively_empty(v, treat_zero_as_empty: bool) -> bool:
    if v is None:
        return True
    if isinstance(v, str):
        return v.strip() == ""
    if isinstance(v, bool):
        return False
    if isinstance(v, (int, float)):
        return abs(float(v)) < 1e-12 if treat_zero_as_empty else False
    return False

def select_columns_to_keep(ws, data_rows: List[int], c_left: int, c_right: int, cfg: RenderConfig) -> List[int]:
    cols = list(range(c_left, c_right + 1))
    if not cfg.hide_empty_columns:
        return cols
    kept = []
    for c in cols:
        has_any = False
        for r in data_rows:
            if not is_effectively_empty(ws.cell(r, c).value, cfg.treat_zero_as_empty):
                has_any = True
                break
        if has_any:
            kept.append(c)
    return kept if kept else [c_left]


# =========================
# SIZE
# =========================
def excel_colwidth_to_px(width: Optional[float], scale: int, min_px: int=60, max_px: int=560) -> int:
    if width is None:
        width = 8.43
    try:
        w = float(width)
    except Exception:
        w = 8.43
    if w <= 0:
        w = 8.43
    px = int((w * 7 + 5) * scale)
    px = max(px, min_px * scale)
    px = min(px, max_px * scale)
    return px

def points_to_px(points: Optional[float], scale: int, default_pt: float=21.0, min_px: int=28, max_px: int=240) -> int:
    if points is None:
        points = default_pt
    try:
        p = float(points)
    except Exception:
        p = default_pt
    px = int((p * 96.0 / 72.0) * scale)
    px = max(px, int(min_px * scale))
    px = min(px, int(max_px * scale))
    return px

def safe_name(s: str) -> str:
    s = (s or "").strip()
    s = re.sub(r"[^\w\-\.\(\)]+", "_", s, flags=re.UNICODE)
    return s or "sheet"


# =========================
# FONT (no need bundled fonts)
# =========================
@lru_cache(maxsize=1)
def resolve_font_paths() -> Tuple[Optional[str], Optional[str]]:
    env_regular = (os.getenv("FONT_REGULAR_PATH") or "").strip()
    env_bold = (os.getenv("FONT_BOLD_PATH") or "").strip()

    def ok(p: Optional[str]) -> Optional[str]:
        try:
            if p and os.path.exists(p) and os.path.isfile(p):
                return p
        except Exception:
            return None
        return None

    r = ok(env_regular)
    b = ok(env_bold)
    if r and b:
        return r, b
    if r and not b:
        return r, r

    linux_candidates = [
        ("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
         "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
        ("/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
         "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf"),
        ("/usr/share/fonts/truetype/noto/NotoSans-Regular.ttf",
         "/usr/share/fonts/truetype/noto/NotoSans-Bold.ttf"),
        ("/usr/share/fonts/truetype/freefont/FreeSans.ttf",
         "/usr/share/fonts/truetype/freefont/FreeSansBold.ttf"),
    ]
    for reg, bold in linux_candidates:
        rr = ok(reg)
        bb = ok(bold)
        if rr and bb:
            return rr, bb
        if rr and not bb:
            return rr, rr

    win_candidates = [
        ("C:/Windows/Fonts/segoeui.ttf", "C:/Windows/Fonts/segoeuib.ttf"),
        ("C:/Windows/Fonts/arial.ttf",   "C:/Windows/Fonts/arialbd.ttf"),
    ]
    for reg, bold in win_candidates:
        rr = ok(reg)
        bb = ok(bold)
        if rr and bb:
            return rr, bb
        if rr and not bb:
            return rr, rr

    return None, None

@lru_cache(maxsize=512)
def load_font_cached(path: Optional[str], size: int) -> ImageFont.ImageFont:
    try:
        if path and os.path.exists(path):
            return ImageFont.truetype(path, int(size))
    except Exception:
        pass
    return ImageFont.load_default()


# =========================
# MERGED CELLS
# =========================
@dataclass
class MergeSpanAdj:
    r1: int
    r2: int
    c_first: int
    c_last: int
    text_cell_col: int

def build_merge_maps_for_cols(ws, r_top: int, r_bottom: int, cols: List[int]) -> Tuple[Dict[Tuple[int, int], MergeSpanAdj], Set[Tuple[int, int]]]:
    anchors: Dict[Tuple[int, int], MergeSpanAdj] = {}
    covered: Set[Tuple[int, int]] = set()

    for m in ws.merged_cells.ranges:
        r1, c1, r2, c2 = m.min_row, m.min_col, m.max_row, m.max_col
        if r1 < r_top or r2 > r_bottom:
            continue

        kept_in = [c for c in cols if c1 <= c <= c2]
        if not kept_in:
            continue

        anchor_col = kept_in[0]
        anchors[(r1, anchor_col)] = MergeSpanAdj(
            r1=r1, r2=r2, c_first=kept_in[0], c_last=kept_in[-1], text_cell_col=c1
        )

        for rr in range(r1, r2 + 1):
            for cc in kept_in:
                if not (rr == r1 and cc == anchor_col):
                    covered.add((rr, cc))

    return anchors, covered


# =========================
# TEXT WRAP (NO ELLIPSIS)
# =========================
def _text_width(draw: ImageDraw.ImageDraw, text: str, font: ImageFont.ImageFont) -> float:
    try:
        return float(draw.textlength(text, font=font))
    except Exception:
        b = draw.textbbox((0, 0), text, font=font)
        return float(b[2] - b[0])

def _line_height(font: ImageFont.ImageFont) -> int:
    b = font.getbbox("Ag")
    return max(1, int(b[3] - b[1]))

def _split_long_token(draw: ImageDraw.ImageDraw, token: str, font: ImageFont.ImageFont, max_w: int) -> List[str]:
    # bẻ một token dài theo ký tự để không tràn ô
    if not token:
        return [""]
    out = []
    cur = ""
    for ch in token:
        cand = cur + ch
        if _text_width(draw, cand, font) <= max_w or not cur:
            cur = cand
        else:
            out.append(cur)
            cur = ch
    if cur:
        out.append(cur)
    return out

def wrap_text_lines(draw: ImageDraw.ImageDraw, text: str, font: ImageFont.ImageFont, max_w: int) -> List[str]:
    # wrap theo từ; nếu 1 từ quá dài thì bẻ theo ký tự
    s = (text or "").strip()
    if not s:
        return []
    if max_w <= 4:
        return [s]

    words = re.split(r"\s+", s)
    lines: List[str] = []
    line = ""

    for w in words:
        if not w:
            continue

        # nếu w quá dài -> bẻ
        if _text_width(draw, w, font) > max_w:
            pieces = _split_long_token(draw, w, font, max_w)
        else:
            pieces = [w]

        for piece in pieces:
            if not line:
                line = piece
            else:
                cand = line + " " + piece
                if _text_width(draw, cand, font) <= max_w:
                    line = cand
                else:
                    lines.append(line)
                    line = piece

    if line:
        lines.append(line)
    return lines

def draw_text(draw: ImageDraw.ImageDraw, xy, txt: str, font: ImageFont.ImageFont, fill: str, stroke_width: int = 0):
    try:
        if stroke_width and stroke_width > 0:
            draw.text(xy, txt, font=font, fill=fill, stroke_width=stroke_width, stroke_fill=fill)
        else:
            draw.text(xy, txt, font=font, fill=fill)
    except TypeError:
        draw.text(xy, txt, font=font, fill=fill)

def draw_wrapped_in_box(
    draw: ImageDraw.ImageDraw,
    box: Tuple[int, int, int, int],
    text: str,
    font: ImageFont.ImageFont,
    fill: str,
    align: str,
    pad_x: int,
    pad_y: int,
    line_gap: int,
    stroke_width: int
):
    x1, y1, x2, y2 = box
    w = max(1, x2 - x1)
    h = max(1, y2 - y1)

    max_w = max(1, w - 2 * pad_x)
    lines = wrap_text_lines(draw, text, font, max_w) if text else []
    if not lines:
        return

    lh = _line_height(font)
    total_h = len(lines) * lh + (len(lines) - 1) * line_gap
    start_y = y1 + (h - total_h) // 2

    for i, ln in enumerate(lines):
        lw = _text_width(draw, ln, font)
        if align == "center":
            tx = x1 + (w - lw) / 2
        elif align == "right":
            tx = x2 - pad_x - lw
        else:
            tx = x1 + pad_x
        ty = start_y + i * (lh + line_gap)
        draw_text(draw, (tx, ty), ln, font, fill, stroke_width=stroke_width)


# =========================
# RENDER CHUNK
# =========================
def render_chunk_to_png_bytes(ws, header_rows: List[int], data_rows: List[int], cols: List[int], cfg: RenderConfig) -> bytes:
    render_scale = cfg.final_scale * cfg.super_sample

    rows = header_rows + data_rows
    r_top, r_bottom = rows[0], rows[-1]

    regular_path, bold_path = resolve_font_paths()
    title_font  = load_font_cached(bold_path or regular_path, 15 * render_scale)
    header_font = load_font_cached(bold_path or regular_path, 12 * render_scale)
    data_font   = load_font_cached((bold_path if cfg.data_use_bold else regular_path) or bold_path or regular_path, 12 * render_scale)
    footer_font = load_font_cached(bold_path or regular_path, 11 * render_scale)

    # ---- base col widths from Excel
    col_px: List[int] = []
    for c in cols:
        w = ws.column_dimensions[get_column_letter(c)].width
        col_px.append(excel_colwidth_to_px(w, render_scale))

    # ---- base row heights from Excel
    row_px: List[int] = []
    for r in rows:
        h = ws.row_dimensions[r].height
        default_pt = 24.0 if r in header_rows else 21.0
        row_px.append(points_to_px(h, render_scale, default_pt=default_pt))

    # ---- build merge maps first (needed for measuring spans)
    merge_anchors, merge_covered = build_merge_maps_for_cols(ws, r_top, r_bottom, cols)

    col_index = {c: i for i, c in enumerate(cols)}
    row_index = {r: i for i, r in enumerate(rows)}

    pad_x = int(cfg.cell_pad_x_px * render_scale)
    pad_y = int(cfg.cell_pad_y_px * render_scale)
    line_gap = int(cfg.line_gap_px * render_scale)
    widen_cap = int(cfg.col_widen_cap_px * render_scale)

    # ---- measure helper
    tmp = Image.new("RGB", (10, 10), "white")
    mdraw = ImageDraw.Draw(tmp)

    # ---- AUTO WIDEN COLUMNS (để header không bị cụt)
    if cfg.auto_widen_columns:
        # chỉ widen theo cell "không merge" và header text trực tiếp trên cột
        for c in cols:
            ci = col_index[c]
            cur_w = col_px[ci]
            need_w = cur_w

            # header row 1 text (nếu có)
            hr = header_rows[0]
            ht = format_cell_value(ws.cell(hr, c), thousand_sep=".", decimal_sep=",")
            if ht:
                tw = _text_width(mdraw, ht, header_font) + 2 * pad_x + int(10 * render_scale)
                need_w = max(need_w, int(tw))

            # data rows: nới vừa phải (đặc biệt cột tên)
            for r in data_rows:
                vt = ws.cell(r, c).value
                if vt is None:
                    continue
                t = format_cell_value(ws.cell(r, c), thousand_sep=".", decimal_sep=",")
                if not t:
                    continue
                # chỉ dùng để widen nhẹ, vẫn wrap là chính
                tw = _text_width(mdraw, t, data_font) + 2 * pad_x + int(10 * render_scale)
                need_w = max(need_w, int(min(tw, widen_cap)))

            if need_w > cur_w:
                col_px[ci] = min(need_w, widen_cap)

    # ---- AUTO INCREASE ROW HEIGHTS to fit wrapped text (NO CUT)
    def span_box_for_anchor(adj: MergeSpanAdj) -> Tuple[int, int]:
        # returns (width_px, height_px_current)
        c_first_i = col_index[adj.c_first]
        c_last_i  = col_index[adj.c_last]
        w = sum(col_px[c_first_i:c_last_i + 1])

        r1_i = row_index.get(adj.r1, 0)
        r2_i = row_index.get(adj.r2, 0)
        h = sum(row_px[r1_i:r2_i + 1])
        return w, h

    # pass 1: compute required heights
    # IMPORTANT: phải lặp vài vòng vì tăng height làm các merge span thay đổi tổng height
    for _ in range(2):
        for r in rows:
            ri = row_index[r]
            is_header = r in header_rows

            for c in cols:
                if (r, c) in merge_covered:
                    continue

                adj = merge_anchors.get((r, c))
                if adj:
                    # chỉ đo ở anchor (r1, anchor_col)
                    if not (r == adj.r1 and c == adj.c_first):
                        continue

                # determine span width & current height span
                if adj:
                    span_w, span_h = span_box_for_anchor(adj)
                    cell_obj = ws.cell(r, adj.text_cell_col)  # text lấy từ c1
                else:
                    span_w = col_px[col_index[c]]
                    span_h = row_px[ri]
                    cell_obj = ws.cell(r, c)

                txt = format_cell_value(cell_obj, thousand_sep=".", decimal_sep=",")
                if not txt:
                    continue

                font = header_font if is_header else data_font
                stroke = cfg.header_stroke_width if is_header else cfg.data_stroke_width

                max_w = max(1, span_w - 2 * pad_x)
                lines = wrap_text_lines(mdraw, txt, font, max_w)
                if not lines:
                    continue

                lh = _line_height(font)
                needed_h = len(lines) * lh + (len(lines) - 1) * line_gap + 2 * pad_y + int(stroke * 2)

                if adj and adj.r2 > adj.r1:
                    # merged vertical: ensure total height of merged span is enough
                    r1_i = row_index.get(adj.r1, ri)
                    r2_i = row_index.get(adj.r2, ri)
                    cur_total = sum(row_px[r1_i:r2_i + 1])
                    if needed_h > cur_total:
                        extra = needed_h - cur_total
                        per = int(math.ceil(extra / float(r2_i - r1_i + 1)))
                        for k in range(r1_i, r2_i + 1):
                            row_px[k] += per
                else:
                    # normal row: ensure row height enough
                    if needed_h > row_px[ri]:
                        row_px[ri] = needed_h

    # ---- compute final image size (after auto adjust)
    outer_pad = int(12 * render_scale)
    title_h = int(52 * render_scale)
    footer_h = int(46 * render_scale)
    gap_before_footer = int(10 * render_scale)

    grid_w = sum(col_px)
    grid_h = sum(row_px)

    img_w = grid_w + outer_pad * 2
    grid_x0 = outer_pad
    grid_y0 = outer_pad + title_h

    footer_y1 = grid_y0 + grid_h + gap_before_footer
    footer_y2 = footer_y1 + footer_h
    img_h = footer_y2 + outer_pad

    img = Image.new("RGB", (img_w, img_h), "white")
    draw = ImageDraw.Draw(img)

    # ===== Title bar
    title_x1, title_y1 = outer_pad, outer_pad
    title_x2, title_y2 = img_w - outer_pad, outer_pad + title_h
    draw.rounded_rectangle((title_x1, title_y1, title_x2, title_y2), radius=int(10 * render_scale),
                           fill="#ffffff", outline="#e5e7eb", width=max(1, render_scale))
    accent_w = int(5 * render_scale)
    draw.rounded_rectangle((title_x1, title_y1, title_x1 + accent_w, title_y2),
                           radius=int(10 * render_scale), fill="#2563eb", outline=None)

    tx = title_x1 + accent_w + int(12 * render_scale)
    ty = title_y1 + int(14 * render_scale)
    draw_text(draw, (tx, ty), cfg.title_text, title_font, cfg.text_color, stroke_width=1)

    # ===== Grid coords
    xs = [grid_x0]
    for w in col_px:
        xs.append(xs[-1] + w)
    ys = [grid_y0]
    for h in row_px:
        ys.append(ys[-1] + h)

    def draw_cell_rect(x1, y1, x2, y2, bg=None):
        if bg:
            draw.rectangle((x1, y1, x2, y2), fill=bg)
        draw.rectangle((x1, y1, x2, y2), outline=cfg.border_color, width=max(1, render_scale))

    data_row0 = data_rows[0] if data_rows else 0

    for r in rows:
        ri = row_index[r]
        for c in cols:
            if (r, c) in merge_covered:
                continue

            ci = col_index[c]
            x1, x2 = xs[ci], xs[ci + 1]
            y1, y2 = ys[ri], ys[ri + 1]

            adj = merge_anchors.get((r, c))
            if adj:
                c_first_i = col_index[adj.c_first]
                c_last_i  = col_index[adj.c_last]
                x1, x2 = xs[c_first_i], xs[c_last_i + 1]
                r1_i = row_index.get(adj.r1, ri)
                r2_i = row_index.get(adj.r2, ri)
                y1, y2 = ys[r1_i], ys[r2_i + 1]

            is_header = r in header_rows

            # background
            if is_header:
                bg = cfg.header_bg if (r == header_rows[0]) else cfg.header2_bg
            else:
                if cfg.zebra_enabled:
                    parity = (r - data_row0) % 2
                    bg = cfg.zebra_a if parity == 0 else cfg.zebra_b
                else:
                    bg = None

            draw_cell_rect(x1, y1, x2, y2, bg=bg)

            # cell value (NO CUT)
            cell = ws.cell(r, adj.text_cell_col) if adj else ws.cell(r, c)
            txt = format_cell_value(cell, thousand_sep=".", decimal_sep=",")
            if not txt:
                continue

            font = header_font if is_header else data_font
            stroke = cfg.header_stroke_width if is_header else cfg.data_stroke_width

            # alignment
            if is_header:
                align = "center"
            else:
                v = cell.value
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    align = "right"
                elif c == cols[0]:
                    align = "center"
                else:
                    align = "left"

            draw_wrapped_in_box(
                draw=draw,
                box=(x1, y1, x2, y2),
                text=txt,
                font=font,
                fill=cfg.text_color,
                align=align,
                pad_x=pad_x,
                pad_y=pad_y,
                line_gap=line_gap,
                stroke_width=stroke
            )

    # divider under header
    if cfg.draw_header_divider:
        y_header_bottom = ys[len(header_rows)]
        draw.line((xs[0], y_header_bottom, xs[-1], y_header_bottom), fill=cfg.header_divider_color, width=max(1, render_scale))

    # ===== Footer
    fx1, fy1 = outer_pad, footer_y1
    fx2, fy2 = img_w - outer_pad, footer_y2
    draw.rounded_rectangle((fx1, fy1, fx2, fy2), radius=int(10 * render_scale),
                           fill="#f8fafc", outline="#e5e7eb", width=max(1, render_scale))

    # footer wrap center
    footer_padx = int(10 * render_scale)
    footer_pady = int(6 * render_scale)
    draw_wrapped_in_box(
        draw=draw,
        box=(fx1 + footer_padx, fy1 + footer_pady, fx2 - footer_padx, fy2 - footer_pady),
        text=cfg.footer_text,
        font=footer_font,
        fill=cfg.text_color,
        align="center",
        pad_x=0,
        pad_y=0,
        line_gap=int(2 * render_scale),
        stroke_width=1
    )

    # ===== Supersample downscale + sharpen
    if cfg.super_sample > 1:
        out_w = img_w // cfg.super_sample
        out_h = img_h // cfg.super_sample
        img = img.resize((out_w, out_h), resample=Image.Resampling.LANCZOS)
        img = img.filter(ImageFilter.UnsharpMask(radius=1.2, percent=140, threshold=3))

    bio = BytesIO()
    img.save(bio, format="PNG")
    return bio.getvalue()


# =========================
# SPLIT WORKBOOK
# =========================
def split_workbook(xlsx_bytes: bytes, cfg: RenderConfig, target_sheet: Optional[str]) -> List[Tuple[str, bytes]]:
    wb = load_workbook(BytesIO(xlsx_bytes), data_only=True)
    sheets = [target_sheet] if (target_sheet and target_sheet in wb.sheetnames) else wb.sheetnames
    outputs: List[Tuple[str, bytes]] = []

    for sname in sheets:
        ws = wb[sname]

        header_row, header_rows_count, col_stt, col_ma, col_tl = find_header(ws, scan_rows=160)
        header_rows = list(range(header_row, header_row + header_rows_count))
        data_start = header_row + header_rows_count

        total_row = find_total_row(ws, data_start_row=data_start, col_stt=col_stt, c_left_scan=col_stt, c_right_scan=col_tl)
        last_data_row = total_row - 1
        if last_data_row < data_start:
            continue

        # BỎ STT: từ Mã NV -> Tổng lương
        c_left = col_ma
        c_right = col_tl

        total_employees = last_data_row - data_start + 1
        parts = int(math.ceil(total_employees / float(cfg.rows_per_image)))

        for i in range(parts):
            srow = data_start + i * cfg.rows_per_image
            erow = min(srow + cfg.rows_per_image - 1, last_data_row)
            data_rows = list(range(srow, erow + 1))

            cols = select_columns_to_keep(ws, data_rows, c_left, c_right, cfg)

            png = render_chunk_to_png_bytes(ws, header_rows, data_rows, cols, cfg)
            fname = f"{safe_name(sname)}_part{i+1:03d}_rows{srow}-{erow}.png"
            outputs.append((fname, png))

    return outputs


# =========================
# SIDEBAR OPTIONS
# =========================
with st.sidebar:
    st.header("Tuỳ chọn xuất ảnh")

    rows_per_image = st.slider("Số lao động / ảnh", 1, 20, 5, 1)
    final_scale = st.slider("Độ to ảnh (FINAL_SCALE)", 1, 5, 3, 1)
    super_sample = st.slider("Độ nét (SUPER_SAMPLE)", 1, 3, 2, 1)

    st.divider()
    zebra_enabled = st.checkbox("Tô nền đan xen (zebra)", value=True)

    st.divider()
    st.subheader("Cột trống / cột = 0")
    export_empty_zero_cols = st.checkbox("✅ TÍCH để vẫn xuất cột trống & cột toàn = 0", value=False)

    st.divider()
    st.subheader("Chữ dữ liệu đậm")
    data_stroke = st.slider("Độ đậm dữ liệu (stroke)", 0, 4, 2, 1)

    st.divider()
    title_text = st.text_input("Tiêu đề", value="HỢP CHÍ - Nguyễn Huệ HR")
    footer_text = st.text_input("Footer", value="Liên hệ Nguyễn Huệ HR - 0356 227 868 | timvieclam.9phut.com")

    st.divider()
    st.subheader("Màu header (pro)")
    header_bg = st.color_picker("Header chính", value="#eaf2ff")
    header2_bg = st.color_picker("Header phụ", value="#f3f8ff")
    border_color = st.color_picker("Màu viền", value="#cbd5e1")

    st.divider()
    st.subheader("Hiển thị đủ chữ (không …)")
    auto_widen = st.checkbox("Auto nới cột (để header không cụt)", value=True)

# Build cfg
cfg = RenderConfig(
    rows_per_image=rows_per_image,
    final_scale=final_scale,
    super_sample=super_sample,
    zebra_enabled=zebra_enabled,
    title_text=title_text,
    footer_text=footer_text,
    header_bg=header_bg,
    header2_bg=header2_bg,
    border_color=border_color,
    data_stroke_width=data_stroke,
    auto_widen_columns=auto_widen,
)

# checkbox logic: xuất cột trống & 0
if export_empty_zero_cols:
    cfg.hide_empty_columns = False
    cfg.treat_zero_as_empty = False
else:
    cfg.hide_empty_columns = True
    cfg.treat_zero_as_empty = True


# =========================
# MAIN
# =========================
uploaded = st.file_uploader("Tải lên file Excel (.xlsx)", type=["xlsx"])
if not uploaded:
    st.info("Hãy tải file Excel lên để bắt đầu.")
    st.stop()

xlsx_bytes = uploaded.read()

try:
    wb_tmp = load_workbook(BytesIO(xlsx_bytes), data_only=True)
    sheetnames = wb_tmp.sheetnames
except Exception as e:
    st.error(f"Không đọc được file Excel: {e}")
    st.stop()

col1, col2 = st.columns([1, 1])
with col1:
    sheet_choice = st.selectbox("Chọn sheet", options=["(Tất cả)"] + sheetnames)
with col2:
    run_btn = st.button("🚀 Xuất ảnh", use_container_width=True)

if not run_btn:
    st.stop()

target_sheet = None if sheet_choice == "(Tất cả)" else sheet_choice

with st.spinner("Đang xử lý và render ảnh..."):
    try:
        images = split_workbook(xlsx_bytes, cfg, target_sheet)
    except Exception as e:
        st.exception(e)
        st.stop()

if not images:
    st.warning("Không xuất được ảnh nào. Kiểm tra lại: header (STT/Mã NV/Tổng lương) và dòng TỔNG.")
    st.stop()

zip_bio = BytesIO()
with zipfile.ZipFile(zip_bio, "w", compression=zipfile.ZIP_DEFLATED) as zf:
    for fname, data in images:
        zf.writestr(fname, data)
zip_bytes = zip_bio.getvalue()

st.success(f"✅ Xong! Đã tạo {len(images)} ảnh.")
st.download_button(
    "⬇️ Tải ZIP ảnh",
    data=zip_bytes,
    file_name="bang_luong_images.zip",
    mime="application/zip",
    use_container_width=True,
)

st.subheader("Xem trước")
preview_n = min(6, len(images))
cols_prev = st.columns(3)
for i in range(preview_n):
    fname, data = images[i]
    with cols_prev[i % 3]:
        st.image(data, caption=fname, use_container_width=True)
